#!/usr/bin/env python3
"""minimax-xlsx skill: build a high-end .xlsx AWS cost report (openpyxl).

In-container invocation (subprocess, by the reporting tool):
    python skills/minimax_xlsx/build_xlsx.py --spec /path/spec.json
Writes spec["output_path"] and prints the absolute path on stdout.
Also importable for tests: build_xlsx(spec: dict) -> output_path.

Layout:
  * "Overview" - title, KPI card band, a derived-insight panel and a spacious
    VISUAL BREAKDOWN. When chart PNGs are supplied (spec["charts"]) they are
    embedded full-width and generously spaced (same high-res charts as the PDF).
    With no PNGs, two clean NATIVE (editable) Excel charts are generated instead.
  * "Details" - the line-item table with a Share column, in-cell data bars,
    a frozen header and an auto-filter.

Spec schema (output_path + rows required):
{
  "title": str, "subtitle": str,
  "period":   {"start": "YYYY-MM-DD", "end": "YYYY-MM-DD"},
  "currency": {"display": "IDR", "usd_rate": float, "as_of": str},
  "rows":     [{"service": str, "usd": float, "display": float?}],
  "total":    {"usd": float, "display": float?},
  "charts":   ["/abs/chart.png", ...],   # embedded full-width when present
  "notes":    [str, ...],
  "output_path": "/abs/out.xlsx"
}
"""
from __future__ import annotations

import argparse
import datetime
import json
import os
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import report_common as rc  # noqa: E402

FONT = "Calibri"
EMBED_W = 760          # px width for embedded chart images
ROW_PX = 20            # approx px per default row (for anchor spacing)


def _hex(c: str) -> str:
    return c.lstrip("#").upper()


def _font(size=10, bold=False, color=rc.BODY, name=FONT):
    return Font(name=name, size=size, bold=bold, color=_hex(color))


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=_hex(color))


_ACCENT_MED = Side(style="medium", color=_hex(rc.ACCENT))


def _merge_set(ws, rng, value=None, font=None, align=None, fill=None):
    ws.merge_cells(rng)
    tl = ws[rng.split(":")[0]]
    if value is not None:
        tl.value = value
    if font:
        tl.font = font
    if align:
        tl.alignment = align
    if fill:
        for row in ws[rng]:
            for cell in row:
                cell.fill = fill
    return tl


def _build_details(wb, spec, rows, total, display, show_display):
    ws = wb.create_sheet("Details")
    ws.sheet_view.showGridLines = False

    headers = ["Service", "Amount (USD)"]
    if show_display:
        headers.append(f"Amount ({display})")
    headers.append("Share")
    ncol = len(headers)
    share_col = ncol
    usd_col = 2

    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _fill(rc.ACCENT)
        cell.font = _font(10, bold=True, color="#FFFFFF")
        cell.alignment = left if c == 1 else right
        cell.border = Border(bottom=Side(style="thin", color=_hex(rc.ACCENT_DK)))

    tot_usd = rc._f((total or {}).get("usd")) or sum(rc._f(r.get("usd")) for r in rows) or 1e-9
    r = 2
    for i, row in enumerate(rows):
        usd = rc._f(row.get("usd"))
        band = _fill(rc.ZEBRA) if i % 2 else _fill("#FFFFFF")
        sc = ws.cell(row=r, column=1, value=row.get("service"))
        sc.font = _font(10, color=rc.BODY)
        sc.alignment = left
        c2 = ws.cell(row=r, column=usd_col, value=round(usd, 2))
        c2.number_format = "#,##0.00"
        c2.font = _font(10, color=rc.INK)
        c2.alignment = right
        col = 3
        if show_display:
            dv = row.get("display")
            cd = ws.cell(row=r, column=col, value=(rc._f(dv) if dv is not None else None))
            cd.number_format = "#,##0"
            cd.font = _font(10, color=rc.BODY)
            cd.alignment = right
            col += 1
        cs = ws.cell(row=r, column=share_col, value=(usd / tot_usd))
        cs.number_format = "0.0%"
        cs.font = _font(10, color=rc.MUTED)
        cs.alignment = right
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = band
            cell.border = Border(bottom=Side(style="thin", color=_hex("#EEF1F5")))
        r += 1

    last_data = r - 1
    tcell = ws.cell(row=r, column=1, value="Total")
    tcell.font = _font(10, bold=True, color=rc.INK)
    tcell.alignment = left
    tc2 = ws.cell(row=r, column=usd_col, value=round(rc._f((total or {}).get("usd") or tot_usd), 2))
    tc2.number_format = "#,##0.00"
    tc2.font = _font(10, bold=True, color=rc.INK)
    tc2.alignment = right
    col = 3
    if show_display:
        dv = (total or {}).get("display")
        td = ws.cell(row=r, column=col, value=(rc._f(dv) if dv is not None else None))
        td.number_format = "#,##0"
        td.font = _font(10, bold=True, color=rc.INK)
        td.alignment = right
        col += 1
    tsh = ws.cell(row=r, column=share_col, value=1.0)
    tsh.number_format = "0%"
    tsh.font = _font(10, bold=True, color=rc.INK)
    tsh.alignment = right
    for c in range(1, ncol + 1):
        ws.cell(row=r, column=c).border = Border(top=Side(style="medium", color=_hex(rc.INK)))
    total_row = r

    ws.column_dimensions["A"].width = 44
    ws.column_dimensions[get_column_letter(usd_col)].width = 16
    if show_display:
        ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(share_col)].width = 12

    if last_data >= 2:
        bar_rng = f"{get_column_letter(usd_col)}2:{get_column_letter(usd_col)}{last_data}"
        ws.conditional_formatting.add(bar_rng, DataBarRule(
            start_type="num", start_value=0, end_type="max",
            color=_hex(rc.ACCENT), showValue=True))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncol)}{max(1, last_data)}"
    ws.sheet_properties.tabColor = _hex(rc.MUTED)
    return dict(ws=ws, ncol=ncol, first_data=2, last_data=last_data,
                total_row=total_row, usd_col=usd_col, service_col=1)


def _embed_images(ov, paths, row):
    """Embed chart PNGs full-width, generously spaced (returns next free row)."""
    try:
        from PIL import Image as PImage
    except Exception:  # noqa: BLE001
        PImage = None
    r = row
    for p in paths:
        img = XLImage(p)
        w0, h0 = img.width or EMBED_W, img.height or int(EMBED_W * 0.56)
        if PImage:
            try:
                with PImage.open(p) as im:
                    w0, h0 = im.size
            except Exception:  # noqa: BLE001
                pass
        scale = EMBED_W / float(w0 or EMBED_W)
        img.width = EMBED_W
        img.height = int((h0 or EMBED_W) * scale)
        ov.add_image(img, f"B{r}")
        r += int(img.height / ROW_PX) + 4
    return r


def _native_charts(ov, det, row):
    """Fallback when no chart PNGs are supplied: two clean editable charts."""
    ws = det["ws"]
    first, last = det["first_data"], det["last_data"]
    if last < first:
        return
    cats = Reference(ws, min_col=det["service_col"], min_row=first, max_row=last)
    vals = Reference(ws, min_col=det["usd_col"], min_row=1, max_row=last)
    ramp = [_hex(c) for c in rc.CHART_RAMP]

    dough = DoughnutChart()
    dough.title = "Cost composition"
    dough.add_data(vals, titles_from_data=True)
    dough.set_categories(cats)
    dough.holeSize = 60
    dough.height = 9.2
    dough.width = 13.5
    dl = DataLabelList()
    dl.showPercent = True
    dl.showCatName = False
    dl.showSerName = False
    dl.showVal = False
    dl.showLegendKey = False
    dl.numFmt = "0%"
    dough.dataLabels = dl
    if dough.legend is not None:
        dough.legend.position = "r"
        dough.legend.overlay = False
    if dough.series:
        s = dough.series[0]
        for i in range(last - first + 1):
            dp = DataPoint(idx=i)
            dp.graphicalProperties.solidFill = ramp[i % len(ramp)]
            dp.graphicalProperties.line.solidFill = "FFFFFF"
            dp.graphicalProperties.line.width = 19050
            s.data_points.append(dp)
    ov.add_chart(dough, f"B{row}")

    bar = BarChart()
    bar.type = "bar"
    bar.title = "Cost by service (USD)"
    bar.add_data(vals, titles_from_data=True)
    bar.set_categories(cats)
    bar.legend = None
    bar.height = 9.2
    bar.width = 15.0
    bar.gapWidth = 55
    if bar.series:
        bar.series[0].graphicalProperties.solidFill = _hex(rc.ACCENT)
        bar.series[0].graphicalProperties.line.noFill = True
    ov.add_chart(bar, f"B{row + 20}")


def _build_overview(ov, det, spec, metrics, total, display, show_display):
    ov.sheet_view.showGridLines = False
    ov.sheet_properties.tabColor = _hex(rc.ACCENT)

    for col in ("A", "D", "G", "J", "M"):
        ov.column_dimensions[col].width = 2.4
    for col in ("B", "C", "E", "F", "H", "I", "K", "L"):
        ov.column_dimensions[col].width = 10.6

    _merge_set(ov, "B1:L1", spec.get("title") or "AWS Cost Report",
               _font(20, bold=True, color=rc.INK), Alignment(vertical="center"))
    ov.row_dimensions[1].height = 28
    if spec.get("subtitle"):
        _merge_set(ov, "B2:L2", str(spec["subtitle"]), _font(11.5, color=rc.MUTED))

    cur = spec.get("currency") or {}
    period = spec.get("period") or {}
    chips = []
    if period.get("start"):
        chips.append(f"Period {period['start']} \u2192 {period.get('end', '')}")
    if cur.get("usd_rate"):
        chips.append(f"FX 1 USD = {rc.money(cur['usd_rate'], display, 2)} {display}")
    chips.append("Generated " + datetime.datetime.now(datetime.timezone.utc).strftime("%d %b %Y"))
    _merge_set(ov, "B3:L3", "   \u00b7   ".join(chips), _font(8.5, color=rc.FAINT))

    cards = rc.kpi_cards(metrics, total, display, show_display)
    for card, (c1, c2) in zip(cards, [("B", "C"), ("E", "F"), ("H", "I"), ("K", "L")]):
        accent = bool(card.get("accent"))
        ground = rc.ACCENT_TINT if accent else rc.PANEL
        _merge_set(ov, f"{c1}5:{c2}5", str(card["label"]).upper(),
                   _font(8, bold=True, color=rc.MUTED), Alignment(vertical="center"), _fill(ground))
        _merge_set(ov, f"{c1}6:{c2}6", str(card["value"]),
                   _font(float(card.get("size", 18)), bold=True,
                         color=rc.ACCENT_DK if accent else rc.INK),
                   Alignment(vertical="center"), _fill(ground))
        _merge_set(ov, f"{c1}7:{c2}7", str(card.get("sub", "")),
                   _font(8.5, color=rc.MUTED), Alignment(vertical="center"), _fill(ground))
    ov.row_dimensions[6].height = 26

    row = 9
    bullets = metrics.get("bullets") or []
    if bullets:
        _merge_set(ov, f"B{row}:L{row}", "KEY INSIGHTS",
                   _font(8, bold=True, color=rc.ACCENT_DK), Alignment(vertical="center"),
                   _fill(rc.ACCENT_TINT))
        row += 1
        for b in bullets:
            tl = _merge_set(ov, f"B{row}:L{row}", "\u25b8  " + b,
                            _font(10, color=rc.BODY),
                            Alignment(vertical="center", indent=1), _fill(rc.ACCENT_TINT))
            tl.border = Border(left=_ACCENT_MED)
            row += 1
        row += 2

    _merge_set(ov, f"B{row}:L{row}", "VISUAL BREAKDOWN", _font(8, bold=True, color=rc.ACCENT))
    row += 1
    paths = [p for p in (spec.get("charts") or []) if p and os.path.isfile(p)]
    if paths:
        _embed_images(ov, paths, row)
    else:
        _native_charts(ov, det, row)


def build_xlsx(spec: dict) -> str:
    cur = spec.get("currency") or {}
    display = cur.get("display") or "USD"
    show_display = display.upper() != "USD"
    rate = cur.get("usd_rate")
    rows = rc.sorted_rows(spec.get("rows"))
    total = spec.get("total") or {}
    metrics = rc.compute_insights(rows, total, display, show_display, rate)

    wb = Workbook()
    ov = wb.active
    ov.title = "Overview"
    det = _build_details(wb, spec, rows, total, display, show_display)
    _build_overview(ov, det, spec, metrics, total, display, show_display)

    notes = list(spec.get("notes") or [])
    if not any("Cost Explorer" in str(n) for n in notes):
        notes.append("Figures from AWS Cost Explorer (UnblendedCost).")
    ws = det["ws"]
    nr = det["total_row"] + 2
    ws.cell(row=nr, column=1, value="Notes").font = _font(9, bold=True, color=rc.INK)
    nr += 1
    for n in notes:
        ws.cell(row=nr, column=1, value=f"\u2013 {n}").font = _font(8.5, color=rc.MUTED)
        nr += 1

    out = spec["output_path"]
    os.makedirs(os.path.dirname(os.path.abspath(out)), exist_ok=True)
    wb.save(out)
    return os.path.abspath(out)


def main():
    ap = argparse.ArgumentParser(description="Build an .xlsx cost report")
    ap.add_argument("--spec", required=True, help="path to spec JSON")
    args = ap.parse_args()
    with open(args.spec, encoding="utf-8") as f:
        spec = json.load(f)
    print(build_xlsx(spec))


if __name__ == "__main__":
    main()
