#!/usr/bin/env python3
"""Skill demo + xlsx round-trip (Task 7 acceptance).

Builds a sample .xlsx and .pdf via the two skills, then reads the .xlsx back to
verify integrity. Usage: python skills/demo.py [out_dir]
"""
import importlib.util
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))


def _load(rel, name):
    spec = importlib.util.spec_from_file_location(name, os.path.join(HERE, rel))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def sample_spec(out_path, charts):
    rate = 17958.44
    rows = [
        {"service": "Claude Opus 4.8 (Amazon Bedrock Edition)", "usd": 112.66},
        {"service": "Tax", "usd": 14.74},
        {"service": "Claude Sonnet 4.6 (Amazon Bedrock Edition)", "usd": 11.90},
        {"service": "Amazon S3", "usd": 0.13},
    ]
    for r in rows:
        r["display"] = round(r["usd"] * rate)
    total_usd = round(sum(r["usd"] for r in rows), 2)
    return {
        "title": "AWS Cost Report",
        "subtitle": "demo-account - June 2026",
        "period": {"start": "2026-06-01", "end": "2026-07-01"},
        "currency": {"display": "IDR", "usd_rate": rate, "as_of": "Sun, 19 Jul 2026 00:02:31 +0000"},
        "rows": rows,
        "total": {"usd": total_usd, "display": round(total_usd * rate)},
        "charts": charts,
        "notes": ["Figures from AWS Cost Explorer (UnblendedCost).", "Demo report."],
        "output_path": out_path,
    }


def main():
    out_dir = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        __import__("tempfile").gettempdir(), "cba-skill-demo")
    os.makedirs(out_dir, exist_ok=True)

    bx = _load("minimax_xlsx/build_xlsx.py", "build_xlsx")
    bp = _load("minimax_pdf/build_pdf.py", "build_pdf")

    xlsx_out = os.path.join(out_dir, "report.xlsx")
    pdf_out = os.path.join(out_dir, "report.pdf")

    x = bx.build_xlsx(sample_spec(xlsx_out, []))
    p = bp.build_pdf(sample_spec(pdf_out, []))
    print("xlsx:", x, os.path.getsize(x), "bytes")
    print("pdf :", p, os.path.getsize(p), "bytes")

    # xlsx round-trip
    from openpyxl import load_workbook
    wb = load_workbook(x)
    ws = wb["Cost Report"]
    assert ws["A1"].value == "AWS Cost Report", ws["A1"].value
    # find a numeric cost cell
    found = any(isinstance(ws.cell(row=r, column=2).value, (int, float))
                for r in range(6, 12))
    assert found, "no numeric cost cells found on round-trip"

    # pdf sanity
    with open(p, "rb") as f:
        assert f.read(5) == b"%PDF-", "not a PDF"

    print("SKILL DEMO OK (xlsx round-trip verified)")


if __name__ == "__main__":
    main()
