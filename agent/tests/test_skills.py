"""Offline unit tests for the vendored report skills (openpyxl + reportlab)."""
import base64
import importlib.util
import os
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TINY_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg==")


def _load(rel, name):
    spec = importlib.util.spec_from_file_location(name, os.path.join(_ROOT, rel))
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


bx = _load("skills/minimax_xlsx/build_xlsx.py", "build_xlsx_mod")
bp = _load("skills/minimax_pdf/build_pdf.py", "build_pdf_mod")


def _spec(out, charts):
    return {
        "title": "AWS Cost Report",
        "subtitle": "prod - June 2026",
        "period": {"start": "2026-06-01", "end": "2026-07-01"},
        "currency": {"display": "IDR", "usd_rate": 17958.44, "as_of": "Sun, 19 Jul 2026"},
        "rows": [
            {"service": "Amazon EC2", "usd": 800.0, "display": round(800.0 * 17958.44)},
            {"service": "Amazon S3", "usd": 434.56, "display": round(434.56 * 17958.44)},
        ],
        "total": {"usd": 1234.56, "display": round(1234.56 * 17958.44)},
        "charts": charts,
        "notes": ["Figures from AWS Cost Explorer (UnblendedCost)."],
        "output_path": out,
    }


def test_xlsx_build_and_roundtrip(tmp_path):
    chart = tmp_path / "c.png"
    chart.write_bytes(TINY_PNG)
    out = str(tmp_path / "r.xlsx")
    path = bx.build_xlsx(_spec(out, [str(chart)]))
    assert os.path.isfile(path)
    wb = load_workbook(path)
    assert "Overview" in wb.sheetnames and "Details" in wb.sheetnames
    ov, det = wb["Overview"], wb["Details"]
    # title lives on the Overview sheet
    assert any(str(c.value or "").strip() == "AWS Cost Report"
               for row in ov.iter_rows() for c in row)
    # a positive numeric cost cell in the Details table (column B)
    vals = [det.cell(row=r, column=2).value for r in range(2, 8)]
    assert any(isinstance(v, (int, float)) and v > 0 for v in vals)
    # the supplied chart image is embedded (full-width) on the Overview
    with zipfile.ZipFile(path) as z:
        assert any(n.startswith("xl/media/image") for n in z.namelist())


def test_pdf_build_valid(tmp_path):
    chart = tmp_path / "c.png"
    chart.write_bytes(TINY_PNG)
    out = str(tmp_path / "r.pdf")
    path = bp.build_pdf(_spec(out, [str(chart)]))
    assert os.path.isfile(path)
    with open(path, "rb") as f:
        assert f.read(5) == b"%PDF-"
    reader = PdfReader(path)
    assert len(reader.pages) >= 1


def test_xlsx_usd_only_no_display_column(tmp_path):
    out = str(tmp_path / "usd.xlsx")
    spec = _spec(out, [])
    spec["currency"] = {"display": "USD"}
    path = bx.build_xlsx(spec)
    wb = load_workbook(path)
    det = wb["Details"]
    header = [det.cell(row=1, column=c).value for c in range(1, 5)]
    assert header[0] == "Service" and header[1] == "Amount (USD)"
    assert header[2] == "Share"        # Share directly after USD - no display column
    assert header[3] is None
    # no display-currency column anywhere in the header
    assert not any(str(h or "").startswith("Amount (") and "USD" not in str(h) for h in header)
    # no PNGs supplied -> two native (editable) charts are generated as a fallback
    with zipfile.ZipFile(path) as z:
        assert sum(n.startswith("xl/charts/chart") for n in z.namelist()) >= 2
